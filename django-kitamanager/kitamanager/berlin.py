from openpyxl import load_workbook
from efc.interfaces.iopenpyxl import OpenpyxlInterface
import datetime
from typing import Optional, Dict
from decimal import Decimal

"""
Read "Senatsabrechungen" for Berlin/Germany .xlsx files
See https://www.berlin.de/sen/jugend/traegerservice/isbj/dokumentation_abrechnungsuebersicht_.pdf
"""


class BerlinInvoice:
    def __init__(self, filename):
        self._filename = filename
        self._wb = load_workbook(filename=filename)
        self._interface = OpenpyxlInterface(wb=self._wb, use_cache=True)

    @property
    def wb(self):
        """the workbook"""
        return self._wb

    @property
    def date(self) -> datetime.date:
        ws = self.wb["Abrechnungsübersicht"]
        # older (than 2025-07) senatsabrechnungen have the date in the cell D28, newer in I18
        value = ws["D28"].value
        if not value:
            value = ws["I18"].value

        return datetime.datetime.strptime(value, "%m/%y").date()

    @property
    def pay(self) -> Optional[Decimal]:
        """
        Pay amount in Euro
        """
        ws = self.wb["Abrechnungsübersicht"]
        for cell in ws["A"]:
            if cell.value == "Summe:":
                # older (than 2025-07) senatsabrechnungen have the summe in the column D, newer in I
                value = self._interface.calc_cell(f"D{cell.row}", "Abrechnungsübersicht")
                if not value:
                    value = self._interface.calc_cell(f"I{cell.row}", "Abrechnungsübersicht")
                return Decimal(value).quantize(Decimal("0.00"))
        raise Exception("Unable to find Summe from Abrechnungsuebersicht")

    @property
    def children(self) -> Dict[str, Dict[str, str]]:
        """
        List of children from the invoice
        """
        children: Dict[str, Dict[str, str]] = dict()
        ws = self.wb["Vertragsübersicht"]
        # die kinderliste faengt bei Reihe 8 an
        for row in ws.iter_rows(min_row=9):
            # ende when kein gutscheincode da ist
            if not row[4] or not row[4].value:
                break
            data = dict()
            name = row[5].value.split(",")
            data["last_name"] = name[0].strip()
            data["first_name"] = name[1].strip()
            pay_tags = []
            # betreuungsumfang
            if row[13].value == "erweitert":
                pay_tags.append("ganztag erweitert")
            elif row[13].value == "ganztags":
                pay_tags.append("ganztag")
            elif row[13].value == "teilzeit":
                pay_tags.append("teilzeit")
            elif row[13].value == "halbtag":
                pay_tags.append("halbtag")
            else:
                raise Exception(f'Unknown Betreuungsumfang "{row[13].value}" in Senatsabrechnung')

            # QM (Quartiersmanagment)
            if row[8].value == "nein":
                pass
            elif row[8].value == "ja":
                pay_tags.append("qm/mss")
            else:
                raise Exception(f'Unknown QM "{row[8].value}" in Senatsabrechnung')

            # MSS (Monitoring Soziale Stadtentwicklung)
            if row[9].value == "nein":
                pass
            elif row[9].value == "ja":
                pay_tags.append("qm/mss")
            else:
                raise Exception(f'Unknown MSS "{row[9].value}" in Senatsabrechnung')

            # Hs (Herkunftssprache)
            if row[10].value == "D":
                pass
            elif row[10].value == "ND":
                pay_tags.append("ndh")
            else:
                raise Exception(f'Unknown Hs "{row[10].value}" in Senatsabrechnung')

            #  SpH (Sozialpäd. Hilfe)
            if row[11].value == "N":
                pass
            elif row[11].value == "A":
                pay_tags.append("integration a")
            elif row[11].value == "B":
                pay_tags.append("integration b")
            else:
                raise Exception(f'Unknown SpH "{row[11].value}" in Senatsabrechnung')

            data["pay_tags"] = sorted(list(set(pay_tags)))
            # voucher number as key
            children[row[4].value] = data

        return children

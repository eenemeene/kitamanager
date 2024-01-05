import os
import datetime
from openpyxl import load_workbook
from typing import Dict
from decimal import Decimal


class BankAccountEntryImport:
    def __init__(self, filename):
        # filename is something like "Kontostände 01.10.2022.xlsx"
        self._filename = filename
        self._wb = load_workbook(filename=filename)

    @property
    def wb(self):
        """the workbook"""
        return self._wb

    @property
    def date(self):
        filename = self._filename
        # self._filename might be a InMemoryUploadedFile
        if getattr(self._filename, "name", None):
            filename = self._filename.name
        return datetime.datetime.strptime(os.path.basename(filename).split(" ")[1].rstrip(".xslx"), "%d.%m.%Y")

    @property
    def balance(self) -> Dict[str, Decimal]:
        ws = self.wb["Tabelle1"]
        balance = dict()
        # row 'C' contains the bank names, row 'T' the values
        for cell in ws["C"]:
            if cell.value and cell.value[0].isdigit():
                name = cell.value.split("·")[1].strip()
                value = Decimal(ws[f"T{cell.row}"].value).quantize(Decimal("0.00"))
                balance[name] = value
        return balance
